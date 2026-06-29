#!/usr/bin/env python3
"""원천 엑셀(학력_자격증_*.xlsm) → 앱 시드 CSV(Data/*.csv) 변환 ETL.

운영부서가 관리하는 마스터 워크북을 quali-fit이 읽는 시드 CSV로 1회 변환한다.
시드 형식은 db.py 의 SEED_PLAN 과 1:1로 맞춘다(헤더명·FK 순서 동일).

== 단계 ==
이 스크립트는 6개 시드 CSV 를 모두 채운다(1·2·3단계 완료):
    01_직원.csv               ← '개인DB' 시트
    02_자격증_마스터.csv       ← '자격증DB' 시트 (+ '자격증' 시트에만 있는 보유자격증 union)
    03_업무코드_마스터.csv      ← '업무별 코드분류표' 시트 (leaf 코드 = WORK-CA-111 ...)
    04_직원_학력.csv           ← '학력' 시트
    05_직원_자격증.csv         ← '자격증' 시트(취득/등록/유효기간) → 만료 모니터링
    06_업무코드_자격증_매핑.csv ← 자격증 카테고리(CA-110 등)를 업무 leaf 로 전개 → 추천 점수화

== 사용법 (repo 루트에서) ==
    SRC="/path/to/학력_자격증_26.05.14_v13.xlsm" python scripts/etl_xlsm_to_seeds.py
    # 또는
    python scripts/etl_xlsm_to_seeds.py "/path/to/학력_자격증.xlsm"
    # 그 뒤 DB 적재:
    .venv/Scripts/python -c "import db; db.seed_from_csv()"

== 주의 ==
- 산출되는 Data/*.csv 는 PII/회사 기밀이다. .gitignore 로 제외되며 절대 커밋 금지.
- 이 스크립트(코드)만 버전관리한다.
- 의존성: openpyxl (엑셀 읽기). pandas 불필요(stdlib csv 사용).
"""
import csv
import os
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "Data"

# --- 시트명 (원천 워크북) ---
SHEET_PERSON = "개인DB"
SHEET_CERT = "자격증DB"
SHEET_EDU = "학력"
SHEET_EMPCERT = "자격증"
SHEET_WORK = "업무별 코드분류표"

# 업무/카테고리 코드 패턴 (예: CA-110, WORK-CA-111 의 접두 제외부)
CODE_RE = re.compile(r"^[A-Z]{2}-\d{3}$")
WORK_PREFIX = "WORK-"   # work_code PK 관례 (issue #28, export.py 가 표시 때 strip)

# --- SEED_PLAN 과 동일한 CSV 헤더 (순서 포함) ---
H_EMPLOYEE = ["직원번호", "이름", "소속", "직책"]
H_CERT = ["코드", "자격증명", "대분류", "중분류", "원가산정검증활용", "자격증내용",
          "수행가능업무", "영향력", "자격유형", "증빙유형", "자격등급구분", "키워드",
          "관련부처", "시행/발급기관"]
H_EDU = ["직원번호", "학력정보번호", "학력", "학위", "학교명", "학부(과)", "전공", "비고"]
H_WORK = ["업무분류코드", "대분류", "중분류", "소분류", "업무구분", "분류기준",
          "관리부서", "책임자", "적용된키워드", "분류근거및설명", "관련지침", "관련법령"]
H_EMP_CERT = ["직원번호", "자격증코드", "취득일", "등록일", "유효기간"]
H_MAP = ["업무분류코드", "자격증코드", "업무관련영향력"]


def s(v) -> str:
    """셀 값을 시드용 문자열로 정규화한다(날짜는 YYYY-MM-DD)."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def importance_to_influence(raw: str) -> str:
    """'중요도' 셀 → influence(1~5 정수 문자열). 비면 중립값 3.

    cert_master.influence 는 INTEGER + CHECK(BETWEEN 1 AND 5) 이라 빈 문자열을
    넣으면 시드가 실패한다. 따라서 항상 유효 정수를 보장한다.
    """
    raw = (raw or "").strip()
    if not raw:
        return "3"
    m = re.search(r"-?\d+", raw)
    if m:
        return str(max(1, min(5, int(m.group()))))
    return {"상": "5", "중": "3", "하": "1",
            "매우높음": "5", "높음": "4", "보통": "3", "낮음": "2"}.get(raw, "3")


def write_csv(name: str, header: list[str], rows: list[list[str]]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / name
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"  wrote {name:28} rows={len(rows)}")


def load_employees(ws):
    """개인DB → employee 행 + 이름→직원번호 매핑. (헤더 2줄, 데이터 3행~)"""
    rows, name_to_id, dups = [], {}, []
    n = 0
    for r in ws.iter_rows(min_row=3, values_only=True):
        name = s(r[1])          # B: 이름
        if not name:
            continue
        dept, title = s(r[2]), s(r[3])   # C: 소속, D: 직책
        n += 1
        emp_id = f"EMP-{n:03d}"
        rows.append([emp_id, name, dept, title])
        if name in name_to_id:
            dups.append(name)
        else:
            name_to_id[name] = emp_id
    return rows, name_to_id, dups


def load_certs(ws):
    """자격증DB → cert_master 행 + 자격증명→코드 매핑 + 코드→카테고리그룹코드.

    (헤더 2줄, 데이터 3행~)  카테고리 그룹코드(CA-110 등)는 KEY(col8)와
    제조(CA)~갈등조정(col9~17) 셀에서 모은다 → 3단계 work_code_cert_map 용.
    """
    rows, name_to_code, dups, categories = [], {}, [], {}
    n = 0
    for r in ws.iter_rows(min_row=3, values_only=True):
        cert_name = s(r[1])     # B: 자격증명
        if not cert_name:
            continue
        ministry = s(r[2])      # C: 관련부처
        issuer = s(r[3])        # D: 시행/발급기관
        license_type = s(r[5])  # F: 구분 (민간/국가/외국...)
        evidence = s(r[6])      # G: 비고 (자격증/등록증)
        importance = s(r[7])    # H: 중요도
        key = s(r[8])           # I: KEY (대표 카테고리 코드, 예: CB-210)
        n += 1
        code = f"CERT-{n:03d}"
        rows.append([
            code, cert_name,
            key, "",                       # 대분류(l1)=KEY, 중분류(l2) 미정
            "", "", "",                    # 원가산정검증활용/자격증내용/수행가능업무
            importance_to_influence(importance),
            license_type, evidence, "", "",  # 자격유형/증빙유형/자격등급구분/키워드
            ministry, issuer,
        ])
        # 카테고리 그룹코드 수집(KEY + CA~CG 등 col8~17)
        groups = {g for g in (s(r[c]) for c in range(8, 18) if c < len(r))
                  if CODE_RE.match(g)}
        if cert_name in name_to_code:
            dups.append(cert_name)
        else:
            name_to_code[cert_name] = code
            categories[code] = groups
    return rows, name_to_code, dups, categories


def load_education(ws, name_to_id):
    """학력 → education 행. 이름으로 직원번호 조인. (헤더 1줄, 데이터 2행~)"""
    rows, missing = [], []
    n = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = s(r[0])          # 이름
        level = s(r[3])         # 학력
        if not name and not level:
            continue
        emp_id = name_to_id.get(name)
        if not emp_id:
            missing.append(name or "(빈 이름)")
            continue
        n += 1
        rows.append([
            emp_id, f"EDU-{n:03d}",
            level,                # 학력 → level
            s(r[4]),              # 학위 → degree
            s(r[5]),              # 학교명 → school
            s(r[6]),              # 학부(과) → faculty
            s(r[7]),              # 전공 → major
            s(r[8]),              # 비고 → note
        ])
    return rows, missing


def load_employee_certs(ws, name_to_id, cert_rows, name_to_code):
    """자격증 시트 → employee_cert 행. (헤더 1줄, 데이터 2행~)

    이름→직원번호, 자격증명→코드로 조인한다. 자격증DB(catalog)에 없는 보유
    자격증은 cert_master(cert_rows)에 union 으로 추가해 FK(RESTRICT)를 충족한다.
    복합 PK(employee_id, cert_code) 충돌은 첫 건만 남기고 경고한다.
    컬럼: 0 이름, 3 자격증, 4 구분, 5 관련부처, 6 시행/발급기관, 7 비고,
          8 취득일, 9 등록일, 10 유효기간, 13 KEY
    """
    rows, miss_name, dups, added = [], [], [], []
    seen = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = s(r[0])
        cert_name = s(r[3])
        if not name and not cert_name:
            continue
        emp_id = name_to_id.get(name)
        if not emp_id:
            miss_name.append(name or "(빈 이름)")
            continue
        if not cert_name:
            continue
        code = name_to_code.get(cert_name)
        if not code:                       # catalog 에 없는 보유자격증 → union 추가
            code = f"CERT-{len(cert_rows) + 1:03d}"
            name_to_code[cert_name] = code
            cert_rows.append([
                code, cert_name,
                s(r[13]), "",              # 대분류=KEY, 중분류
                "", "", "",                # 원가산정검증활용/자격증내용/수행가능업무
                "3",                       # 영향력 중립값
                s(r[4]), s(r[7]), "", "",  # 자격유형=구분, 증빙유형=비고
                s(r[5]), s(r[6]),          # 관련부처, 시행/발급기관
            ])
            added.append(cert_name)
        key = (emp_id, code)
        if key in seen:
            dups.append((name, cert_name))
            continue
        seen.add(key)
        rows.append([emp_id, code, s(r[8]), s(r[9]), s(r[10])])   # 취득/등록/유효기간
    return rows, miss_name, dups, added


def load_work_codes(ws):
    """업무별 코드분류표 → work_code_master 행 + 그룹코드→leaf work_code 매핑.

    4단계 계층(병합셀): col0 대분류 / col1 카테고리(제조(CA)) / col2 중분류 /
    col3 세부분류 / col4 그룹코드(CA-110) / col5 산정·검증명 / col6 leaf코드(CA-111).
    leaf(col6)마다 한 행. work_code = 'WORK-'+leaf. task_type 은 산정/검증.
    그룹코드(col4)→[leaf work_code...] 매핑은 자격증 카테고리 전개에 쓴다.
    (헤더 row4, 데이터 row5~. 병합셀은 forward-fill)
    """
    FILL = [0, 1, 2, 3, 4, 7, 8, 9, 10, 11]   # 병합되어 비는 컬럼 (forward-fill 대상)
    last = {}
    rows, group_to_leaves, bad = [], {}, []
    for r in ws.iter_rows(min_row=5, values_only=True):
        for c in FILL:
            v = s(r[c]) if c < len(r) else ""
            if v:
                last[c] = v
        leaf = s(r[6]) if len(r) > 6 else ""
        if not CODE_RE.match(leaf):
            continue
        c5 = s(r[5]) if len(r) > 5 else ""
        if "산정" in c5:
            task = "산정"
        elif "검증" in c5:
            task = "검증"
        else:                                  # 폴백: leaf 끝자리 홀=산정/짝=검증
            task = "산정" if int(leaf[-1]) % 2 else "검증"
        work_code = WORK_PREFIX + leaf
        rows.append([
            work_code,
            last.get(0, ""),                   # 대분류 → l1
            last.get(2, ""),                   # 중분류 → l2
            last.get(3, ""),                   # 세부분류 → l3(소분류)
            task,                              # 업무구분 (산정/검증)
            last.get(7, ""),                   # 분류 기준
            last.get(8, ""),                   # 관리부서
            last.get(9, ""),                   # 책임자
            "",                                # 적용된키워드 (원천 없음)
            "",                                # 분류근거및설명 (원천 없음)
            last.get(10, ""),                  # 관련지침
            last.get(11, ""),                  # 관련법령
        ])
        group = last.get(4, "")
        if CODE_RE.match(group):
            group_to_leaves.setdefault(group, []).append(work_code)
        else:
            bad.append(leaf)
    return rows, group_to_leaves, bad


def load_work_cert_map(cert_rows, cert_categories, group_to_leaves):
    """자격증 카테고리 그룹코드를 leaf work_code 로 전개해 work_code_cert_map 생성.

    cert 의 카테고리(CA-110 등) → 그 그룹의 leaf(산정·검증 둘 다) 로 펼친다.
    influence 는 해당 cert 의 영향력(cert_master)을 사용. 복합 PK(work,cert) 중복 제거.
    """
    influence_of = {row[0]: row[7] for row in cert_rows}   # cert_code → 영향력
    rows, seen, unmapped = [], set(), set()
    for cert_code, groups in cert_categories.items():
        for g in groups:
            leaves = group_to_leaves.get(g)
            if not leaves:
                unmapped.add(g)
                continue
            for work_code in leaves:
                key = (work_code, cert_code)
                if key in seen:
                    continue
                seen.add(key)
                rows.append([work_code, cert_code, influence_of.get(cert_code, "3")])
    return rows, sorted(unmapped)


def main() -> None:
    src = os.environ.get("SRC") or (sys.argv[1] if len(sys.argv) > 1 else "")
    if not src:
        sys.exit("원천 워크북 경로가 필요합니다. SRC 환경변수 또는 첫 인자로 .xlsm 경로를 주세요.")
    src_path = Path(os.path.expanduser(src))
    if not src_path.exists():
        sys.exit(f"파일을 찾을 수 없습니다: {src_path}")

    print(f"읽는 중: {src_path}")
    wb = openpyxl.load_workbook(src_path, data_only=True, read_only=True)
    for need in (SHEET_PERSON, SHEET_CERT, SHEET_EDU, SHEET_EMPCERT, SHEET_WORK):
        if need not in wb.sheetnames:
            sys.exit(f"필요한 시트가 없습니다: {need!r} (있는 시트: {wb.sheetnames})")

    emp_rows, name_to_id, emp_dups = load_employees(wb[SHEET_PERSON])
    cert_rows, cert_map, cert_dups, cert_categories = load_certs(wb[SHEET_CERT])
    # 보유자격증 처리 시 cert_rows/cert_map 에 catalog 누락분이 union 으로 추가됨 → 02 보다 먼저 실행
    ec_rows, ec_miss, ec_dups, ec_added = load_employee_certs(
        wb[SHEET_EMPCERT], name_to_id, cert_rows, cert_map)
    edu_rows, edu_missing = load_education(wb[SHEET_EDU], name_to_id)
    work_rows, group_to_leaves, work_bad = load_work_codes(wb[SHEET_WORK])
    map_rows, map_unmapped = load_work_cert_map(cert_rows, cert_categories, group_to_leaves)

    print("\n시드 CSV 생성 (Data/):")
    write_csv("01_직원.csv", H_EMPLOYEE, emp_rows)
    write_csv("02_자격증_마스터.csv", H_CERT, cert_rows)   # 자격증DB + 보유자격증 union
    write_csv("03_업무코드_마스터.csv", H_WORK, work_rows)
    write_csv("04_직원_학력.csv", H_EDU, edu_rows)
    write_csv("05_직원_자격증.csv", H_EMP_CERT, ec_rows)
    write_csv("06_업무코드_자격증_매핑.csv", H_MAP, map_rows)

    # --- 경고/리포트 ---
    print("\n요약:")
    print(f"  직원 {len(emp_rows)}명 · 자격증마스터 {len(cert_rows)}종"
          f"(보유 union +{len(ec_added)}) · 학력 {len(edu_rows)}건"
          f" · 보유자격증 {len(ec_rows)}건")
    print(f"  업무코드 {len(work_rows)}개(그룹 {len(group_to_leaves)}) · 업무–자격증 매핑 {len(map_rows)}건")
    if emp_dups:
        print(f"  [경고] 개인DB 동명이인(첫 항목만 매핑됨): {sorted(set(emp_dups))}")
    if cert_dups:
        print(f"  [경고] 자격증DB 중복 자격증명: {sorted(set(cert_dups))}")
    if edu_missing:
        print(f"  [경고] 학력 {len(edu_missing)}건이 개인DB에 없는 이름 → 제외: {sorted(set(edu_missing))}")
    if ec_added:
        print(f"  [정보] catalog 누락 보유자격증 {len(ec_added)}종을 cert_master 에 추가: {sorted(set(ec_added))}")
    if ec_miss:
        print(f"  [경고] 보유자격증 {len(ec_miss)}건이 개인DB에 없는 이름 → 제외: {sorted(set(ec_miss))}")
    if ec_dups:
        print(f"  [경고] (직원,자격증) 중복 {len(ec_dups)}건 → 첫 건만: {ec_dups[:10]}")
    if work_bad:
        print(f"  [경고] 그룹코드 없는 leaf {len(work_bad)}개: {work_bad[:10]}")
    if map_unmapped:
        print(f"  [경고] 업무표에 없는 자격증 카테고리코드 {len(map_unmapped)}: {map_unmapped}")
    print("\n다음: repo 루트에서  python -c \"import db; db.seed_from_csv()\"  로 적재")


if __name__ == "__main__":
    main()
